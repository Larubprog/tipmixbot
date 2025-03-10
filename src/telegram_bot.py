import json
import requests
import os
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Side, Font
from datetime import datetime
import dropbox
from openpyxl.utils import get_column_letter

TELEGRAM_BOT_TOKEN = '8019071023:AAG9JHIzc9GFlkUZlnPtlZf7bCNJSgQiKSw'
TELEGRAM_CHAT_ID = '-4617933873'
DROPBOX_ACCESS_TOKEN = 'sl.u.AFnVj6WdUkqcBE_5TOLaAy8r2yPw_4MB4w6nvEAqbgQcjGvjhClZ1611TSCsIdJ9yd2aVsSG7A3R7l4i3HJxP3lZUtCTOVO_6x830gOfaOMcU34VSyEjjjaPUIWAVGEhtajjvNEXeaOZ9-dhV3_t6lj-BRx0BOY9BxYz-b8XcduFnMizH_VitibeJ7ARTZXCgEoufetgXGvbjM8D4UqOf07ODIOcGIRzpGfgHFvDBXXvqFi5RQQzqY4hF9VtpU6j2f-0_PaBIpS2SYqCjMYpdGewICDHzS7pfYhmVVxV_aokYx_TzomtEdD9Amm_Lm0wm2MzAx8C4U0hYJM64ld2ZwuFuMgXjR5iIX2-dAMogCAB4Y57qnMhU1Dx_zSVhwWvYVwusUDQJTIOWTaEIhex9vGjRsHZC3dhTSd7jrY8uQfZe7YQ_z8mQEIcO8WGDNybJNgaoTbiqwmxcn6Z0TYfA3iaeKjeeotkYuMFCb2o5qupbXBll73I-dacYzZy5B2VIvfgXYqSg0wnif_Ce5lJViaRBmt2pAYHbY-KakHSb1c0KKOQBCXozj5tkfNo1lXs6D8Edg_QOM6cy33l2qoY3e20dyRU6JIBoVB6Drl-8e17pnh1Rr2NYpyHvFqsD3v89cY3z4saC8zJxUlABIMSX89uXuODcStFpRm-NiHYCy-Q1enW-u2DYEkMs625HkSrlzGp4dozdfTskmHF1XW7fxwwVl2wUoEcrHk6qqAHiu3KFgYqFtwfMfFoKP4rvCnYLxRJ9pI9n8RwJEpNd_nQT6Yn3HeqlWFRECEcICCZcQr6Ce9Q5Mjbo1bpBmTMHaj_XXckc2noJAe9zfAiQ8NxmP8fs89QWMMTUY2LqcBiy6Mtn4wFkIIeLlkyzJvARC8Mh09TYNC1ploJ70gYK83SLbu7_JazHbOhYORjeGBJ83Ecp4nCjJNQW0Tq6NxzPEVNpCgqLCXOCVsLWxR9VqLB_6iBtsnW9HcnV_eD7ithaUebWZ5n_b2QC87C_y_0pebHMwgRqYta4lfDG8sj9_jvzpGqj2_3aTQsHjYj-Q09AS0xZZdHHV6Z94_E4eQPWRYNaMHc4bUUf5KN-wDOlOrusqbsw9rKkSphidTrNrv2i_ajFyxVjIn_te1udT-rOAw5_2MujRAlo7Wqzx5phoCqy33Z7kEXRmLr5w6xxgxlrBalkyLMKT7DAkQ2D4Sqi5Eh_b6xC4uINbFGdomoBd8621u17B5KJIasCM1t8neLCun2J7Z00dQqY8mNO-R2kT8r98GFVcXyRnb0Dvlt7gpzcQHvF7Ghl1zQ3BLSxuIW7jQ-HNO45kMukb8glGINB7hmyuPx13L6oD0rCpo8eU5jqLpfKAqEq4aqYx0fPrx66IYC29ag2nFkyUrBI7KHe2a6VDMC_dZq9BkccKiFtp5KwYZGprJVYHmZo-kRUWlpqPzbMBFFjG3M0Zxv91XZF2ZkMeo'

WIN_PROBABILITY_THRESHOLD = 25
WIN_ODD = 1.1
GOAL_PROBABILITY_THRESHOLD = 10
GOAL_ODD = 1.1
DRAW_PROBABILITY_THRESHOLD = 25
DRAW_ODD = 1.1

def send_telegram_message(message):
    url = f"https://api.telegram.org/bot{TELEGRAM_BOT_TOKEN}/sendMessage"
    payload = {
        "chat_id": TELEGRAM_CHAT_ID,
        "text": message,
        "parse_mode": "Markdown"
    }
    response = requests.post(url, json=payload)
    return response.json()

def load_json(file_path):
    with open(file_path, 'r') as file:
        return json.load(file)

def generate_file_name(player1, player2):
    current_time = datetime.now().strftime("%Y%m%d_%H%M%S")
    return f"data/{player1}_vs_{player2}_{current_time}.xlsx"

def generate_excel(data, output_file):
    wb = Workbook()
    ws = wb.active
    ws.title = "Player Stats"
    bold_font = Font(bold=True)
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thick_border = Border(
        left=Side(style="medium"),
        right=Side(style="medium"),
        top=Side(style="medium"),
        bottom=Side(style="medium")
    )
    ws.cell(row=5, column=4).border = thick_border
    ws.cell(row=6, column=4).border = thick_border
    ws.cell(row=7, column=4).border = thick_border
    ws.cell(row=8, column=4).border = thick_border
    ws.cell(row=4, column=5).border = thick_border
    ws.cell(row=5, column=5).border = thick_border
    ws.cell(row=6, column=5).border = thick_border
    ws.cell(row=7, column=5).border = thick_border
    ws.cell(row=8, column=5).border = thick_border
    ws.cell(row=4, column=6).border = thick_border
    ws.cell(row=5, column=6).border = thick_border
    ws.cell(row=6, column=6).border = thick_border
    ws.cell(row=7, column=6).border = thick_border
    ws.cell(row=8, column=6).border = thick_border
    ws.cell(row=4, column=7).border = thick_border
    ws.cell(row=5, column=7).border = thick_border
    ws.cell(row=6, column=7).border = thick_border
    ws.cell(row=7, column=7).border = thick_border
    ws.cell(row=8, column=7).border = thick_border
    ws.cell(row=4, column=11).border = thick_border
    ws.cell(row=4, column=12).border = thick_border
    ws.cell(row=4, column=13).border = thick_border
    ws.cell(row=4, column=14).border = thick_border
    ws.cell(row=4, column=15).border = thick_border
    ws.cell(row=4, column=16).border = thick_border
    ws.cell(row=4, column=17).border = thick_border
    ws.cell(row=5, column=10).border = thick_border
    ws.cell(row=5, column=11).border = thick_border
    ws.cell(row=5, column=12).border = thick_border
    ws.cell(row=5, column=13).border = thick_border
    ws.cell(row=5, column=14).border = thick_border
    ws.cell(row=5, column=15).border = thick_border
    ws.cell(row=5, column=16).border = thick_border
    ws.cell(row=5, column=17).border = thick_border
    ws.cell(row=12, column=4).border = thick_border
    ws.cell(row=13, column=4).border = thick_border
    ws.cell(row=14, column=4).border = thick_border
    ws.cell(row=15, column=4).border = thick_border
    ws.cell(row=11, column=5).border = thick_border
    ws.cell(row=12, column=5).border = thick_border
    ws.cell(row=13, column=5).border = thick_border
    ws.cell(row=14, column=5).border = thick_border
    ws.cell(row=15, column=5).border = thick_border
    ws.cell(row=11, column=6).border = thick_border
    ws.cell(row=12, column=6).border = thick_border
    ws.cell(row=13, column=6).border = thick_border
    ws.cell(row=14, column=6).border = thick_border
    ws.cell(row=15, column=6).border = thick_border
    ws.cell(row=11, column=7).border = thick_border
    ws.cell(row=12, column=7).border = thick_border
    ws.cell(row=13, column=7).border = thick_border
    ws.cell(row=14, column=7).border = thick_border
    ws.cell(row=15, column=7).border = thick_border
    ws.cell(row=11, column=11).border = thick_border
    ws.cell(row=11, column=12).border = thick_border
    ws.cell(row=11, column=13).border = thick_border
    ws.cell(row=11, column=14).border = thick_border
    ws.cell(row=11, column=15).border = thick_border
    ws.cell(row=11, column=16).border = thick_border
    ws.cell(row=11, column=17).border = thick_border
    ws.cell(row=12, column=10).border = thick_border
    ws.cell(row=12, column=11).border = thick_border
    ws.cell(row=12, column=12).border = thick_border
    ws.cell(row=12, column=13).border = thick_border
    ws.cell(row=12, column=14).border = thick_border
    ws.cell(row=12, column=15).border = thick_border
    ws.cell(row=12, column=16).border = thick_border
    ws.cell(row=12, column=17).border = thick_border
    ws.cell(row=19, column=4).border = thick_border
    ws.cell(row=20, column=4).border = thick_border
    ws.cell(row=21, column=4).border = thick_border
    ws.cell(row=22, column=4).border = thick_border
    ws.cell(row=18, column=5).border = thick_border
    ws.cell(row=19, column=5).border = thick_border
    ws.cell(row=20, column=5).border = thick_border
    ws.cell(row=21, column=5).border = thick_border
    ws.cell(row=22, column=5).border = thick_border
    ws.cell(row=18, column=6).border = thick_border
    ws.cell(row=19, column=6).border = thick_border
    ws.cell(row=20, column=6).border = thick_border
    ws.cell(row=21, column=6).border = thick_border
    ws.cell(row=22, column=6).border = thick_border
    ws.cell(row=18, column=7).border = thick_border
    ws.cell(row=19, column=7).border = thick_border
    ws.cell(row=20, column=7).border = thick_border
    ws.cell(row=21, column=7).border = thick_border
    ws.cell(row=22, column=7).border = thick_border
    ws.cell(row=18, column=11).border = thick_border
    ws.cell(row=18, column=12).border = thick_border
    ws.cell(row=18, column=13).border = thick_border
    ws.cell(row=18, column=14).border = thick_border
    ws.cell(row=18, column=15).border = thick_border
    ws.cell(row=18, column=16).border = thick_border
    ws.cell(row=18, column=17).border = thick_border
    ws.cell(row=19, column=10).border = thick_border
    ws.cell(row=19, column=11).border = thick_border
    ws.cell(row=19, column=12).border = thick_border
    ws.cell(row=19, column=13).border = thick_border
    ws.cell(row=19, column=14).border = thick_border
    ws.cell(row=19, column=15).border = thick_border
    ws.cell(row=19, column=16).border = thick_border
    ws.cell(row=19, column=17).border = thick_border
    ws.cell(row=26, column=4).border = thick_border
    ws.cell(row=27, column=4).border = thick_border
    ws.cell(row=28, column=4).border = thick_border
    ws.cell(row=29, column=4).border = thick_border
    ws.cell(row=25, column=5).border = thick_border
    ws.cell(row=26, column=5).border = thick_border
    ws.cell(row=27, column=5).border = thick_border
    ws.cell(row=28, column=5).border = thick_border
    ws.cell(row=29, column=5).border = thick_border
    ws.cell(row=25, column=6).border = thick_border
    ws.cell(row=26, column=6).border = thick_border
    ws.cell(row=27, column=6).border = thick_border
    ws.cell(row=28, column=6).border = thick_border
    ws.cell(row=29, column=6).border = thick_border
    ws.cell(row=25, column=7).border = thick_border
    ws.cell(row=26, column=7).border = thick_border
    ws.cell(row=27, column=7).border = thick_border
    ws.cell(row=28, column=7).border = thick_border
    ws.cell(row=29, column=7).border = thick_border
    ws.cell(row=25, column=11).border = thick_border
    ws.cell(row=25, column=12).border = thick_border
    ws.cell(row=25, column=13).border = thick_border
    ws.cell(row=25, column=14).border = thick_border
    ws.cell(row=25, column=15).border = thick_border
    ws.cell(row=25, column=16).border = thick_border
    ws.cell(row=25, column=17).border = thick_border
    ws.cell(row=26, column=10).border = thick_border
    ws.cell(row=26, column=11).border = thick_border
    ws.cell(row=26, column=12).border = thick_border
    ws.cell(row=26, column=13).border = thick_border
    ws.cell(row=26, column=14).border = thick_border
    ws.cell(row=26, column=15).border = thick_border
    ws.cell(row=26, column=16).border = thick_border
    ws.cell(row=26, column=17).border = thick_border
    player1_name = data.get("player1")
    player2_name = data.get("player2")
    p25_win_player1 = data["stats"]["past_25"]["win_draw_loss"]["win"]
    p25_draw = data["stats"]["past_25"]["win_draw_loss"]["draw"]
    p25_loss_player2 = data["stats"]["past_25"]["win_draw_loss"]["loss"]
    p25_goals_combined = data["stats"]["past_25"]["avg_goals"]["total_goals"]["full_time"]
    p25_goals_1st_player1 = data["stats"]["past_25"]["avg_goals"][player1_name]["first_half"]
    p25_goals_2nd_player1 = data["stats"]["past_25"]["avg_goals"][player1_name]["second_half"]
    p25_goals_1st_player2 = data["stats"]["past_25"]["avg_goals"][player2_name]["first_half"]
    p25_goals_2nd_player2 = data["stats"]["past_25"]["avg_goals"][player2_name]["second_half"]
    p50_win_player1 = data["stats"]["past_50"]["win_draw_loss"]["win"]
    p50_draw = data["stats"]["past_50"]["win_draw_loss"]["draw"]
    p50_loss_player2 = data["stats"]["past_50"]["win_draw_loss"]["loss"]
    p50_goals_combined = data["stats"]["past_50"]["avg_goals"]["total_goals"]["full_time"]
    p50_goals_1st_player1 = data["stats"]["past_50"]["avg_goals"][player1_name]["first_half"]
    p50_goals_2nd_player1 = data["stats"]["past_50"]["avg_goals"][player1_name]["second_half"]
    p50_goals_1st_player2 = data["stats"]["past_50"]["avg_goals"][player2_name]["first_half"]
    p50_goals_2nd_player2 = data["stats"]["past_50"]["avg_goals"][player2_name]["second_half"]
    p30_win_player1 = data["stats"]["past_30_days"]["win_draw_loss"]["win"]
    p30_draw = data["stats"]["past_30_days"]["win_draw_loss"]["draw"]
    p30_loss_player2 = data["stats"]["past_30_days"]["win_draw_loss"]["loss"]
    p30_goals_combined = data["stats"]["past_30_days"]["avg_goals"]["total_goals"]["full_time"]
    p30_goals_1st_player1 = data["stats"]["past_30_days"]["avg_goals"][player1_name]["first_half"]
    p30_goals_2nd_player1 = data["stats"]["past_30_days"]["avg_goals"][player1_name]["second_half"]
    p30_goals_1st_player2 = data["stats"]["past_30_days"]["avg_goals"][player2_name]["first_half"]
    p30_goals_2nd_player2 = data["stats"]["past_30_days"]["avg_goals"][player2_name]["second_half"]
    p25_thresholds = data["stats"]["past_25"]["goal_thresholds"]
    p50_thresholds = data["stats"]["past_50"]["goal_thresholds"]
    p30_thresholds = data["stats"]["past_30_days"]["goal_thresholds"]
    ws.cell(row=3, column=3).value = "Past 25 games"
    ws.cell(row=3, column=3).font = bold_font
    ws.cell(row=3, column=3).alignment = center_align
    ws.cell(row=4, column=5).value = player1_name
    ws.cell(row=4, column=5).font = bold_font
    ws.cell(row=3, column=3).alignment = center_align
    ws.cell(row=4, column=6).value = "Draw"
    ws.cell(row=4, column=6).font = bold_font
    ws.cell(row=4, column=6).alignment = center_align
    ws.cell(row=4, column=7).value = player2_name
    ws.cell(row=4, column=7).font = bold_font
    ws.cell(row=4, column=7).alignment = center_align
    ws.cell(row=5, column=4).value = "Win Percentage"
    ws.cell(row=5, column=4).font = bold_font
    ws.cell(row=5, column=4).alignment = center_align
    ws.cell(row=5, column=5).value = p25_win_player1 / 100
    ws.cell(row=5, column=5).number_format = '0%'
    ws.cell(row=5, column=6).value = p25_draw / 100
    ws.cell(row=5, column=6).number_format = '0%'
    ws.cell(row=5, column=7).value = p25_loss_player2 / 100
    ws.cell(row=5, column=7).number_format = '0%'
    ws.cell(row=6, column=4).value = "Avarage Goals"
    ws.cell(row=6, column=4).font = bold_font
    ws.cell(row=6, column=4).alignment = center_align
    ws.cell(row=6, column=5).value = p25_goals_1st_player1 + p25_goals_2nd_player1
    ws.cell(row=6, column=6).value = p25_goals_combined
    ws.cell(row=6, column=7).value = p25_goals_2nd_player2 + p25_goals_2nd_player2
    ws.cell(row=7, column=4).value = "Avarage Goals 1st half"
    ws.cell(row=7, column=4).font = bold_font
    ws.cell(row=7, column=4).alignment = center_align
    ws.cell(row=7, column=5).value = p25_goals_1st_player1
    ws.cell(row=7, column=6).value = p25_goals_1st_player1 + p25_goals_1st_player2
    ws.cell(row=7, column=7).value = p25_goals_1st_player2
    ws.cell(row=8, column=4).value = "Avarage Goals 2nd half"
    ws.cell(row=8, column=4).font = bold_font
    ws.cell(row=8, column=4).alignment = center_align
    ws.cell(row=8, column=5).value = p25_goals_2nd_player1
    ws.cell(row=8, column=6).value = p25_goals_2nd_player1 + p25_goals_2nd_player2
    ws.cell(row=8, column=7).value = p25_goals_2nd_player2
    ws.cell(row=5, column=10).value = "Total"
    ws.cell(row=5, column=10).font = bold_font
    ws.cell(row=5, column=10).alignment = center_align
    ws.cell(row=4, column=11).value = "2.5"
    ws.cell(row=5, column=11).value = p25_thresholds["above"]["2.5"]
    ws.cell(row=4, column=12).value = "3.5"
    ws.cell(row=5, column=12).value = p25_thresholds["above"]["3.5"]
    ws.cell(row=4, column=13).value = "4.5"
    ws.cell(row=5, column=13).value = p25_thresholds["above"]["4.5"]
    ws.cell(row=4, column=14).value = "5.5"
    ws.cell(row=5, column=14).value = p25_thresholds["above"]["5.5"]
    ws.cell(row=4, column=15).value = "6.5"
    ws.cell(row=5, column=15).value = p25_thresholds["above"]["6.5"]
    ws.cell(row=4, column=16).value = "7.5"
    ws.cell(row=5, column=16).value = p25_thresholds["above"]["7.5"]
    ws.cell(row=4, column=17).value = "8.5"
    ws.cell(row=5, column=17).value = p25_thresholds["above"]["8.5"]
    ws.cell(row=10, column=3).value = "Past 50 games"
    ws.cell(row=10, column=3).font = bold_font
    ws.cell(row=10, column=3).alignment = center_align
    ws.cell(row=11, column=5).value = player1_name
    ws.cell(row=11, column=5).font = bold_font
    ws.cell(row=10, column=3).alignment = center_align
    ws.cell(row=11, column=6).value = "Draw"
    ws.cell(row=11, column=6).font = bold_font
    ws.cell(row=11, column=6).alignment = center_align
    ws.cell(row=11, column=7).value = player2_name
    ws.cell(row=11, column=7).font = bold_font
    ws.cell(row=11, column=7).alignment = center_align
    ws.cell(row=12, column=4).value = "Win Percentage"
    ws.cell(row=12, column=4).font = bold_font
    ws.cell(row=12, column=4).alignment = center_align
    ws.cell(row=12, column=5).value = p50_win_player1 / 100
    ws.cell(row=12, column=5).number_format = '0%'
    ws.cell(row=12, column=6).value = p50_draw / 100
    ws.cell(row=12, column=6).number_format = '0%'
    ws.cell(row=12, column=7).value = p50_loss_player2 / 100
    ws.cell(row=12, column=7).number_format = '0%'
    ws.cell(row=13, column=4).value = "Avarage Goals"
    ws.cell(row=13, column=4).font = bold_font
    ws.cell(row=13, column=4).alignment = center_align
    ws.cell(row=13, column=5).value = p50_goals_1st_player1 + p50_goals_2nd_player1
    ws.cell(row=13, column=6).value = p50_goals_combined
    ws.cell(row=13, column=7).value = p50_goals_2nd_player2 + p50_goals_2nd_player2
    ws.cell(row=14, column=4).value = "Avarage Goals 1st half"
    ws.cell(row=14, column=4).font = bold_font
    ws.cell(row=14, column=4).alignment = center_align
    ws.cell(row=14, column=5).value = p50_goals_1st_player1
    ws.cell(row=14, column=6).value = p50_goals_2nd_player1 + p50_goals_2nd_player2
    ws.cell(row=14, column=7).value = p50_goals_1st_player2
    ws.cell(row=15, column=4).value = "Avarage Goals 2nd half"
    ws.cell(row=15, column=4).font = bold_font
    ws.cell(row=15, column=4).alignment = center_align
    ws.cell(row=15, column=5).value = p50_goals_2nd_player1
    ws.cell(row=15, column=6).value = p50_goals_2nd_player1 + p50_goals_2nd_player2
    ws.cell(row=15, column=7).value = p50_goals_2nd_player2
    ws.cell(row=12, column=10).value = "Total"
    ws.cell(row=12, column=10).font = bold_font
    ws.cell(row=12, column=10).alignment = center_align
    ws.cell(row=11, column=11).value = "2.5"
    ws.cell(row=12, column=11).value = p50_thresholds["above"]["2.5"]
    ws.cell(row=11, column=12).value = "3.5"
    ws.cell(row=12, column=12).value = p50_thresholds["above"]["3.5"]
    ws.cell(row=11, column=13).value = "4.5"
    ws.cell(row=12, column=13).value = p50_thresholds["above"]["4.5"]
    ws.cell(row=11, column=14).value = "5.5"
    ws.cell(row=12, column=14).value = p50_thresholds["above"]["5.5"]
    ws.cell(row=11, column=15).value = "6.5"
    ws.cell(row=12, column=15).value = p50_thresholds["above"]["6.5"]
    ws.cell(row=11, column=16).value = "7.5"
    ws.cell(row=12, column=16).value = p50_thresholds["above"]["7.5"]
    ws.cell(row=11, column=17).value = "8.5"
    ws.cell(row=12, column=17).value = p50_thresholds["above"]["8.5"]
    ws.cell(row=17, column=3).value = "Past 30 days games"
    ws.cell(row=17, column=3).font = bold_font
    ws.cell(row=17, column=3).alignment = center_align
    ws.cell(row=18, column=5).value = player1_name
    ws.cell(row=18, column=5).font = bold_font
    ws.cell(row=18, column=3).alignment = center_align
    ws.cell(row=18, column=6).value = "Draw"
    ws.cell(row=18, column=6).font = bold_font
    ws.cell(row=18, column=6).alignment = center_align
    ws.cell(row=18, column=7).value = player2_name
    ws.cell(row=18, column=7).font = bold_font
    ws.cell(row=18, column=7).alignment = center_align
    ws.cell(row=19, column=4).value = "Win Percentage"
    ws.cell(row=19, column=4).font = bold_font
    ws.cell(row=19, column=4).alignment = center_align
    ws.cell(row=19, column=5).value = p30_win_player1 / 100
    ws.cell(row=19, column=5).number_format = '0%'
    ws.cell(row=19, column=6).value = p30_draw / 100
    ws.cell(row=19, column=6).number_format = '0%'
    ws.cell(row=19, column=7).value = p30_loss_player2 / 100
    ws.cell(row=19, column=7).number_format = '0%'
    ws.cell(row=20, column=4).value = "Avarage Goals"
    ws.cell(row=20, column=4).font = bold_font
    ws.cell(row=20, column=4).alignment = center_align
    ws.cell(row=20, column=5).value = p30_goals_1st_player1 + p30_goals_2nd_player1
    ws.cell(row=20, column=6).value = p30_goals_combined
    ws.cell(row=20, column=7).value = p30_goals_2nd_player2 + p30_goals_2nd_player2
    ws.cell(row=21, column=4).value = "Avarage Goals 1st half"
    ws.cell(row=21, column=4).font = bold_font
    ws.cell(row=21, column=4).alignment = center_align
    ws.cell(row=21, column=5).value = p30_goals_1st_player1
    ws.cell(row=21, column=6).value = p30_goals_1st_player1 + p30_goals_1st_player2
    ws.cell(row=21, column=7).value = p30_goals_1st_player2
    ws.cell(row=22, column=4).value = "Avarage Goals 2nd half"
    ws.cell(row=22, column=4).font = bold_font
    ws.cell(row=22, column=4).alignment = center_align
    ws.cell(row=22, column=5).value = p30_goals_2nd_player1
    ws.cell(row=22, column=6).value = p30_goals_2nd_player1 + p30_goals_2nd_player2
    ws.cell(row=22, column=7).value = p30_goals_2nd_player2
    ws.cell(row=19, column=10).value = "Total"
    ws.cell(row=19, column=10).font = bold_font
    ws.cell(row=19, column=10).alignment = center_align
    ws.cell(row=18, column=11).value = "2.5"
    ws.cell(row=19, column=11).value = p30_thresholds["above"]["2.5"]
    ws.cell(row=18, column=12).value = "3.5"
    ws.cell(row=19, column=12).value = p30_thresholds["above"]["3.5"]
    ws.cell(row=18, column=13).value = "4.5"
    ws.cell(row=19, column=13).value = p30_thresholds["above"]["4.5"]
    ws.cell(row=18, column=14).value = "5.5"
    ws.cell(row=19, column=14).value = p30_thresholds["above"]["5.5"]
    ws.cell(row=18, column=15).value = "6.5"
    ws.cell(row=19, column=15).value = p30_thresholds["above"]["6.5"]
    ws.cell(row=18, column=16).value = "7.5"
    ws.cell(row=19, column=16).value = p30_thresholds["above"]["7.5"]
    ws.cell(row=18, column=17).value = "8.5"
    ws.cell(row=19, column=17).value = p30_thresholds["above"]["8.5"]
    ws.cell(row=24, column=3).value = "Avaraged Data"
    ws.cell(row=24, column=3).font = bold_font
    ws.cell(row=24, column=3).alignment = center_align
    ws.cell(row=25, column=5).value = player1_name
    ws.cell(row=25, column=5).font = bold_font
    ws.cell(row=25, column=3).alignment = center_align
    ws.cell(row=25, column=6).value = "Draw"
    ws.cell(row=25, column=6).font = bold_font
    ws.cell(row=25, column=6).alignment = center_align
    ws.cell(row=25, column=7).value = player2_name
    ws.cell(row=25, column=7).font = bold_font
    ws.cell(row=25, column=7).alignment = center_align
    ws.cell(row=26, column=4).value = "Win Percentage"
    ws.cell(row=26, column=4).font = bold_font
    ws.cell(row=26, column=4).alignment = center_align
    ws.cell(row=26, column=5).value = ((p25_win_player1 + p30_win_player1 +p30_win_player1) / 3 ) / 100
    ws.cell(row=26, column=5).number_format = '0%'
    ws.cell(row=26, column=6).value = ((p25_draw + p50_draw + p30_draw) / 3 ) / 100
    ws.cell(row=26, column=6).number_format = '0%'
    ws.cell(row=26, column=7).value = ((p30_loss_player2 + p25_loss_player2 + p50_loss_player2) / 3 ) / 100
    ws.cell(row=26, column=7).number_format = '0%'
    ws.cell(row=27, column=4).value = "Avarage Goals"
    ws.cell(row=27, column=4).font = bold_font
    ws.cell(row=27, column=4).alignment = center_align
    ws.cell(row=27, column=5).value = ((p25_goals_1st_player1 + p25_goals_2nd_player1) + (p30_goals_1st_player1 + p30_goals_2nd_player1) + (p50_goals_1st_player1 + p50_goals_2nd_player1)) / 3
    ws.cell(row=27, column=6).value = ((p25_goals_1st_player1 + p25_goals_2nd_player1 + p25_goals_2nd_player2 + p25_goals_2nd_player2) + (p30_goals_1st_player1 + p30_goals_2nd_player1 + p30_goals_2nd_player2 + p30_goals_2nd_player2) + (p50_goals_1st_player1 + p50_goals_2nd_player1 + p50_goals_2nd_player2 + p50_goals_2nd_player2)) / 3
    ws.cell(row=27, column=7).value = ((p25_goals_1st_player2 + p25_goals_2nd_player2) + (p30_goals_1st_player2 + p30_goals_2nd_player2) + (p50_goals_1st_player2 + p50_goals_2nd_player2)) / 3
    ws.cell(row=28, column=4).value = "Avarage Goals 1st half"
    ws.cell(row=28, column=4).font = bold_font
    ws.cell(row=28, column=4).alignment = center_align
    ws.cell(row=28, column=5).value = (p25_goals_1st_player1 + p30_goals_1st_player1 + p50_goals_1st_player1) / 3
    ws.cell(row=28, column=6).value = ((p25_goals_1st_player1 + p25_goals_1st_player2) + (p30_goals_1st_player1 + p30_goals_1st_player2) + (p50_goals_1st_player1 + p50_goals_1st_player2)) / 3
    ws.cell(row=28, column=7).value = (p25_goals_1st_player2 + p30_goals_1st_player2 + p50_goals_1st_player2) / 3 
    ws.cell(row=29, column=4).value = "Avarage Goals 2nd half"
    ws.cell(row=29, column=4).font = bold_font
    ws.cell(row=29, column=4).alignment = center_align
    ws.cell(row=29, column=5).value = (p25_goals_2nd_player1 + p30_goals_2nd_player1 + p50_goals_2nd_player1) / 3
    ws.cell(row=29, column=6).value = ((p25_goals_2nd_player1 + p25_goals_2nd_player2) + (p30_goals_2nd_player1 + p30_goals_2nd_player2) + (p50_goals_2nd_player1 + p50_goals_2nd_player2)) / 3
    ws.cell(row=29, column=7).value = (p25_goals_2nd_player2 + p30_goals_2nd_player2 + p50_goals_2nd_player2) / 3 
    ws.cell(row=26, column=10).value = "Total"
    ws.cell(row=26, column=10).font = bold_font
    ws.cell(row=26, column=10).alignment = center_align
    ws.cell(row=25, column=11).value = "2.5"
    ws.cell(row=26, column=11).value = (p25_thresholds["above"]["2.5"] + p30_thresholds["above"]["2.5"] + p50_thresholds["above"]["2.5"]) / 3
    ws.cell(row=25, column=12).value = "3.5"
    ws.cell(row=26, column=12).value = (p25_thresholds["above"]["3.5"] + p30_thresholds["above"]["3.5"] + p50_thresholds["above"]["3.5"]) / 3
    ws.cell(row=25, column=13).value = "4.5"
    ws.cell(row=26, column=13).value = (p25_thresholds["above"]["4.5"] + p30_thresholds["above"]["4.5"] + p50_thresholds["above"]["4.5"]) / 3
    ws.cell(row=26, column=14).value = "5.5"
    ws.cell(row=26, column=14).value = (p25_thresholds["above"]["5.5"] + p30_thresholds["above"]["5.5"] + p50_thresholds["above"]["5.5"]) / 3
    ws.cell(row=25, column=15).value = "6.5"
    ws.cell(row=26, column=15).value = (p25_thresholds["above"]["6.5"] + p30_thresholds["above"]["6.5"] + p50_thresholds["above"]["6.5"]) / 3
    ws.cell(row=25, column=16).value = "7.5"
    ws.cell(row=26, column=16).value = (p25_thresholds["above"]["7.5"] + p30_thresholds["above"]["7.5"] + p50_thresholds["above"]["7.5"]) / 3
    ws.cell(row=25, column=17).value = "8.5"
    ws.cell(row=26, column=17).value = (p25_thresholds["above"]["8.5"] + p30_thresholds["above"]["8.5"] + p50_thresholds["above"]["8.5"]) / 3
    for column_cells in ws.columns:
        max_length = 0
        col_letter = column_cells[0].column_letter  
        for cell in column_cells:
            try:
                if cell.value:  
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        adjusted_width = max_length + 2  
        ws.column_dimensions[col_letter].width = adjusted_width
        ws.column_dimensions['F'].width = 10 
        ws.column_dimensions['E'].width = 10  
        ws.column_dimensions['G'].width = 10  
    for col in range(11, 17):  
        column_letter = get_column_letter(col)  
        ws.column_dimensions[column_letter].width = 10  
    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell.value, (int, float)):
                if cell.coordinate in ['E5', 'F5', 'G5', 'E12', 'F12', 'G12', 'E19', 'F19', 'G19', 'E26', 'F26', 'G26']:
                    cell.number_format = '0.00%'  
                else:
                    cell.number_format = '0.00'  
    games_sheet = wb.create_sheet(title="All Games")
    bold_font = Font(bold=True)
    center_align = Alignment(horizontal="center")
    thick_border = Border(
        left=Side(style="thick"),
        right=Side(style="thick"),
        top=Side(style="thick"),
        bottom=Side(style="thick"),
    )
    games_sheet.append(["Match ID", "Date", "Player 1", "Team 1", "Player 2", "Team 2", "First Half Score", "Total Goals"])
    for cell in games_sheet[1]:
        cell.font = bold_font
        cell.alignment = center_align
        cell.border = thick_border
    if "games" in data:
        for game in data["games"]:
            match_id = game.get("Match ID", "N/A")
            date = game.get("Date", "N/A")  
            player1 = game.get("Player 1", "N/A")
            team1 = game.get("Team 1", "N/A")
            player2 = game.get("Player 2", "N/A")
            team2 = game.get("Team 2", "N/A")
            first_half_score = f"{game.get('Score player 1 1st half', 'N/A')}-{game.get('Score player 2 1st half', 'N/A')}"
            total_goals = f"{game.get('Score player 1', 'N/A')}-{game.get('Score player 2', 'N/A')}"
            games_sheet.append([match_id, date, player1, team1, player2, team2, first_half_score, total_goals])
    else:
        print("No match history found in stats_data.")
    for column_cells in games_sheet.columns:
        max_length = 0
        col_letter = column_cells[0].column_letter
        for cell in column_cells:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        adjusted_width = max_length + 2
        games_sheet.column_dimensions[col_letter].width = adjusted_width
    wb.save(output_file)

def upload_to_dropbox(file_path, dropbox_path):
    dbx = dropbox.Dropbox(DROPBOX_ACCESS_TOKEN)
    with open(file_path, "rb") as f:
        dbx.files_upload(f.read(), dropbox_path, mode=dropbox.files.WriteMode("overwrite"))
    shared_link = dbx.sharing_create_shared_link_with_settings(dropbox_path).url
    return shared_link.replace("?dl=0", "?dl=1")

def compare_odds_with_stats(games_with_odds):
    try:
        games_data = load_json(games_with_odds)
    except FileNotFoundError as e:
        print(f"Error: File not found - {e}")
        return
    except json.JSONDecodeError as e:
        print(f"Error: Invalid JSON format in file - {e}")
        return

    # Load the tippmixpro_upcoming_games.json file
    try:
        tippmix_games = load_json("data/tippmixpro_upcoming_games.json")
    except FileNotFoundError as e:
        print(f"Error: Tippmix file not found - {e}")
        return
    except json.JSONDecodeError as e:
        print(f"Error: Invalid JSON format in Tippmix file - {e}")
        return

    for game in games_data:
        home_team = game['home']
        away_team = game['away']
        market_data = game['market_data']
        player1 = home_team.split('(')[-1].rstrip(')')
        player2 = away_team.split('(')[-1].rstrip(')')

        # Find the corresponding link from tippmixpro_upcoming_games.json
        match_link = None
        for tippmix_game in tippmix_games['games']:
            if tippmix_game['home'] == home_team and tippmix_game['away'] == away_team:
                match_link = tippmix_game['link']
                break

        stats_file = f"data/{player1}_vs_{player2}_stats.json"
        if not os.path.exists(stats_file):
            print(f"Stats file not found for {player1} vs {player2}. Skipping this match.")
            continue

        try:
            stats_data = load_json(stats_file)
        except FileNotFoundError as e:
            print(f"Error: Stats file not found - {e}")
            continue
        except json.JSONDecodeError as e:
            print(f"Error: Invalid JSON format in stats file - {e}")
            continue

        for market in market_data:
            if market['market_title'] == "1X2 - Rendes játékidő - Full Game":
                for odd in market['odds']:
                    if odd['team'] == home_team:
                        home_odds = float(odd['odds'].replace(',', '.'))
                    elif odd['team'] == away_team:
                        away_odds = float(odd['odds'].replace(',', '.'))
                    elif odd['team'] == "Döntetlen":
                        draw_odds = float(odd['odds'].replace(',', '.'))
                home_implied_prob = (1 / home_odds) * 100
                away_implied_prob = (1 / away_odds) * 100
                draw_implied_prob = (1 / draw_odds) * 100
                past_25 = stats_data['stats']['past_25']
                past_50 = stats_data['stats']['past_50']
                past_30_days = stats_data['stats']['past_30_days']

                if past_25['win_draw_loss']['win'] >= WIN_PROBABILITY_THRESHOLD and home_odds >= WIN_ODD:
                    output_file = generate_file_name(player1, player2)
                    generate_excel(stats_data, output_file)
                    dropbox_path = f"/{output_file}"
                    shareable_link = upload_to_dropbox(output_file, dropbox_path)
                    message = (
                        f"🚨 **Opportunity Detected** 🚨\n\n"  
                        f"**Bet On:** Home Win ({player1})\n\n"  
                        f"**Match:** {player1} vs {player2}\n\n"  
                        f"**Odds:** {home_odds}\n\n"  
                        f"**(Past 25 Games):** {past_25['win_draw_loss']['win']}%\n"
                        f"**(Past 50 Games):** {past_50['win_draw_loss']['win']}%\n"
                        f"**(Past 30 Days):** {past_30_days['win_draw_loss']['win']}%\n\n"  
                        f"**Match Link:** {match_link}\n\n"  # Include the match link
                        f"{shareable_link}\n"
                    )
                    print(message)
                    send_telegram_message(message)

                if past_25['win_draw_loss']['loss'] >= WIN_PROBABILITY_THRESHOLD and away_odds >= WIN_ODD:
                    output_file = generate_file_name(player1, player2)
                    generate_excel(stats_data, output_file)
                    dropbox_path = f"/{output_file}"
                    shareable_link = upload_to_dropbox(output_file, dropbox_path)
                    message = (
                        f"🚨 **Opportunity Detected** 🚨\n\n"  
                        f"**Bet On:** Away Win ({player2})\n\n"  
                        f"**Match:** {player1} vs {player2}\n\n"  
                        f"**Odds:** {away_odds}\n\n"  
                        f"**(Past 25 Games):** {past_25['win_draw_loss']['loss']}%\n"
                        f"**(Past 50 Games):** {past_50['win_draw_loss']['loss']}%\n"
                        f"**(Past 30 Days):** {past_30_days['win_draw_loss']['loss']}%\n\n"  
                        f"**Match Link:** {match_link}\n\n"  # Include the match link
                        f"{shareable_link}\n"
                    )
                    print(message)
                    send_telegram_message(message)

                if past_25['win_draw_loss']['draw'] >= DRAW_PROBABILITY_THRESHOLD and draw_odds >= DRAW_ODD:
                    output_file = generate_file_name(player1, player2)
                    generate_excel(stats_data, output_file)
                    dropbox_path = f"/{output_file}"
                    shareable_link = upload_to_dropbox(output_file, dropbox_path)
                    message = (
                        f"🚨 **Opportunity Detected** 🚨\n\n"  
                        f"**Bet On:** Draw\n\n"  
                        f"**Match:** {player1} vs {player2}\n\n"  
                        f"**Odds:** {draw_odds}\n\n"  
                        f"**(Past 25 Games):** {past_25['win_draw_loss']['draw']}%\n"
                        f"**(Past 50 Games):** {past_50['win_draw_loss']['draw']}%\n"
                        f"**(Past 30 Days):** {past_30_days['win_draw_loss']['draw']}%\n\n"  
                        f"**Match Link:** {match_link}\n\n"  # Include the match link
                        f"{shareable_link}\n"
                    )
                    print(message)
                    send_telegram_message(message)

            if "Gólszám - Rendes játékidő" in market['market_title']:  
                for odd in market['odds']:
                    line = odd.get('line')
                    line_key = str(line).replace(',', '.')
                    over_odds = float(odd.get('over', '0').replace(',', '.'))
                    under_odds = float(odd.get('under', '0').replace(',', '.'))
                    over_implied_prob = (1 / over_odds) * 100 if over_odds > 0 else 0
                    under_implied_prob = (1 / under_odds) * 100 if under_odds > 0 else 0
                    past_25_goal_thresholds = stats_data['stats']['past_25']['goal_thresholds']
                    past_50_goal_thresholds = stats_data['stats']['past_50']['goal_thresholds']
                    past_30_days_goal_thresholds = stats_data['stats']['past_30_days']['goal_thresholds']
                    above_prob_25 = past_25_goal_thresholds['above'].get(line_key, 0)
                    below_prob_25 = past_25_goal_thresholds['below'].get(line_key, 0)
                    above_prob_50 = past_50_goal_thresholds['above'].get(line_key, 0)
                    below_prob_50 = past_50_goal_thresholds['below'].get(line_key, 0)
                    above_prob_30 = past_30_days_goal_thresholds['above'].get(line_key, 0)
                    below_prob_30 = past_30_days_goal_thresholds['below'].get(line_key, 0)

                    if above_prob_25 >= GOAL_PROBABILITY_THRESHOLD and over_odds >= GOAL_ODD:
                        output_file = generate_file_name(player1, player2)
                        generate_excel(stats_data, output_file)
                        dropbox_path = f"/{output_file}"
                        shareable_link = upload_to_dropbox(output_file, dropbox_path)
                        message = (
                            f"🚨 **Opportunity Detected** 🚨\n\n"  
                            f"**Bet On:** Over {line} Goals\n\n"  
                            f"**Match:** {player1} vs {player2}\n\n"  
                            f"**Odds:** {over_odds}\n\n"  
                            f"**(Past 25 Games):** {above_prob_25}%\n"
                            f"**(Past 50 Games):** {above_prob_50}%\n"
                            f"**(Past 30 Days):** {above_prob_30}%\n\n"  
                            f"**Match Link:** {match_link}\n\n"  # Include the match link
                            f"{shareable_link}\n"
                        )
                        print(message)
                        send_telegram_message(message)

                    if below_prob_25 >= GOAL_PROBABILITY_THRESHOLD and under_odds >= GOAL_ODD:
                        output_file = generate_file_name(player1, player2)
                        generate_excel(stats_data, output_file)
                        dropbox_path = f"/{output_file}"
                        shareable_link = upload_to_dropbox(output_file, dropbox_path)
                        message = (
                            f"🚨 **Opportunity Detected** 🚨\n\n"  
                            f"**Bet On:** Under {line} Goals\n\n"  
                            f"**Match:** {player1} vs {player2}\n\n"  
                            f"**Odds:** {under_odds}\n\n"  
                            f"**(Past 25 Games):** {below_prob_25}%\n"
                            f"**(Past 50 Games):** {below_prob_50}%\n"
                            f"**(Past 30 Days):** {below_prob_30}%\n\n"  
                            f"**Match Link:** {match_link}\n\n"  # Include the match link
                            f"{shareable_link}\n"
                        )
                        print(message)
                        send_telegram_message(message)

if __name__ == "__main__":
    games_with_odds_file = "data/games_with_odds.json"
    compare_odds_with_stats(games_with_odds_file)
